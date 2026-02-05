"""Analyse an Amazon bulk XLSX/XLS file in-memory (no DB writes)."""

import logging
import os
from collections import defaultdict

import openpyxl

logger = logging.getLogger(__name__)

# Metric columns shared across sheets
_METRIC_COLS = ("Impressions", "Clicks", "Spend", "Sales", "Orders")


def _safe_float(v, default=0.0):
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def _read_sheet(wb, name):
    """Return (headers, rows) for *name* or (None, []) if missing."""
    if name not in wb.sheetnames:
        return None, []
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None, []
    headers = [str(h).strip() if h else "" for h in rows[0]]
    return headers, rows[1:]


def _open_workbook(filepath: str):
    """
    Open an XLSX or XLS file and return an openpyxl-compatible workbook.
    For .xls files, converts to .xlsx via xlrd first.
    """
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".xls":
        import xlrd
        xls_wb = xlrd.open_workbook(filepath)
        xlsx_wb = openpyxl.Workbook()
        # Remove default sheet
        xlsx_wb.remove(xlsx_wb.active)
        for sheet_name in xls_wb.sheet_names():
            xls_sheet = xls_wb.sheet_by_name(sheet_name)
            ws = xlsx_wb.create_sheet(title=sheet_name)
            for row_idx in range(xls_sheet.nrows):
                for col_idx in range(xls_sheet.ncols):
                    ws.cell(row=row_idx + 1, column=col_idx + 1,
                            value=xls_sheet.cell_value(row_idx, col_idx))
        return xlsx_wb
    return openpyxl.load_workbook(filepath, data_only=True)


def _rows_to_dicts(headers, rows):
    """Convert list-of-tuples to list-of-dicts keyed by header."""
    return [dict(zip(headers, r)) for r in rows]


def _compute_ratios(d):
    """Add CTR / CPC / ACOS / ROAS to a dict that already has the raw sums."""
    imps = d.get("impressions", 0)
    clicks = d.get("clicks", 0)
    spend = d.get("spend", 0)
    sales = d.get("sales", 0)
    d["ctr"] = round(clicks / imps * 100, 2) if imps else 0
    d["cpc"] = round(spend / clicks, 2) if clicks else 0
    d["acos"] = round(spend / sales * 100, 2) if sales else 0
    d["roas"] = round(sales / spend, 2) if spend else 0
    return d


# ── public entry point ──────────────────────────────────────────
def analyze_bulk_file(filepath: str) -> dict:
    wb = _open_workbook(filepath)

    result = {
        "sheets_found": [],
        "overview": {},
        "performance_summary": {},
        "campaigns_table": [],
        "entity_breakdown": [],
        "search_terms_top": [],
        "search_terms_wasted": [],
    }

    # Record which sheets are present
    for sn in wb.sheetnames:
        ws = wb[sn]
        result["sheets_found"].append({"name": sn, "rows": ws.max_row or 0})

    # ── SP Campaigns sheet ──────────────────────────────────────
    sp_headers, sp_rows = _read_sheet(wb, "Sponsored Products Campaigns")
    sp_data = _rows_to_dicts(sp_headers, sp_rows) if sp_headers else []

    # Entity counts
    entity_counts = defaultdict(lambda: {"count": 0, "spend": 0.0})
    for row in sp_data:
        etype = row.get("Entity", "")
        if etype:
            entity_counts[etype]["count"] += 1
            entity_counts[etype]["spend"] += _safe_float(row.get("Spend"))

    result["entity_breakdown"] = sorted(
        [{"entity": k, **v} for k, v in entity_counts.items()],
        key=lambda x: x["spend"],
        reverse=True,
    )

    # Overview counts
    campaigns = [r for r in sp_data if r.get("Entity") == "Campaign"]
    ad_groups = [r for r in sp_data if r.get("Entity") == "Ad Group"]
    keywords = [r for r in sp_data if r.get("Entity") == "Keyword"]
    product_targets = [r for r in sp_data if r.get("Entity") == "Product Targeting"]
    neg_keywords = [r for r in sp_data if r.get("Entity") in
                    ("Negative Keyword", "Campaign Negative Keyword",
                     "Campaign Negative Product Targeting", "Negative Product Targeting")]

    result["overview"] = {
        "campaigns": len(campaigns),
        "ad_groups": len(ad_groups),
        "keywords": len(keywords),
        "product_targets": len(product_targets),
        "negative_keywords": len(neg_keywords),
        "total_rows_sp": len(sp_data),
    }

    # Performance summary (aggregate from campaign-level rows only so we
    # don't double-count spend that appears on sub-entity rows too).
    perf = {"impressions": 0, "clicks": 0, "spend": 0.0, "sales": 0.0, "orders": 0}
    for r in campaigns:
        perf["impressions"] += int(_safe_float(r.get("Impressions")))
        perf["clicks"] += int(_safe_float(r.get("Clicks")))
        perf["spend"] += _safe_float(r.get("Spend"))
        perf["sales"] += _safe_float(r.get("Sales"))
        perf["orders"] += int(_safe_float(r.get("Orders")))
    _compute_ratios(perf)
    result["performance_summary"] = perf

    # Campaigns table
    camp_table = []
    for r in campaigns:
        name = (r.get("Campaign Name") or
                r.get("Campaign Name (Informational only)") or "")
        spend = _safe_float(r.get("Spend"))
        sales = _safe_float(r.get("Sales"))
        clicks = int(_safe_float(r.get("Clicks")))
        imps = int(_safe_float(r.get("Impressions")))
        orders = int(_safe_float(r.get("Orders")))
        camp_table.append({
            "campaign_id": r.get("Campaign ID", ""),
            "portfolio_id": r.get("Portfolio ID", ""),
            "name": name,
            "state": r.get("State", ""),
            "targeting_type": r.get("Targeting Type", ""),
            "daily_budget": _safe_float(r.get("Daily Budget")),
            "impressions": imps,
            "clicks": clicks,
            "spend": round(spend, 2),
            "sales": round(sales, 2),
            "orders": orders,
            "ctr": round(clicks / imps * 100, 2) if imps else 0,
            "cpc": round(spend / clicks, 2) if clicks else 0,
            "acos": round(spend / sales * 100, 2) if sales else 0,
            "roas": round(sales / spend, 2) if spend else 0,
        })
    camp_table.sort(key=lambda x: x["spend"], reverse=True)
    result["campaigns_table"] = camp_table

    # ── Search Term Report ──────────────────────────────────────
    st_headers, st_rows = _read_sheet(wb, "SP Search Term Report")
    st_data = _rows_to_dicts(st_headers, st_rows) if st_headers else []

    # Also try SB Search Term Report
    sb_st_headers, sb_st_rows = _read_sheet(wb, "SB Search Term Report")
    if sb_st_headers:
        st_data += _rows_to_dicts(sb_st_headers, sb_st_rows)

    # Aggregate search terms
    st_agg = defaultdict(lambda: {
        "impressions": 0, "clicks": 0, "spend": 0.0, "sales": 0.0, "orders": 0,
    })
    for r in st_data:
        term = r.get("Customer Search Term", "")
        if not term:
            continue
        st_agg[term]["impressions"] += int(_safe_float(r.get("Impressions")))
        st_agg[term]["clicks"] += int(_safe_float(r.get("Clicks")))
        st_agg[term]["spend"] += _safe_float(r.get("Spend"))
        st_agg[term]["sales"] += _safe_float(r.get("Sales"))
        st_agg[term]["orders"] += int(_safe_float(r.get("Orders")))

    all_terms = []
    for term, m in st_agg.items():
        entry = {"search_term": term, **m}
        _compute_ratios(entry)
        all_terms.append(entry)

    all_terms.sort(key=lambda x: x["spend"], reverse=True)
    result["search_terms_top"] = all_terms[:50]
    result["search_terms_wasted"] = sorted(
        [t for t in all_terms if t["spend"] > 0 and t["orders"] == 0],
        key=lambda x: x["spend"],
        reverse=True,
    )[:50]

    # Add search term overview counts
    result["overview"]["search_terms_total"] = len(all_terms)
    result["overview"]["search_terms_wasted"] = len(
        [t for t in all_terms if t["spend"] > 0 and t["orders"] == 0]
    )

    # ── Data for suggestions engine ─────────────────────────────
    # Targets: keywords + product targeting rows with metrics
    targets = []
    for r in sp_data:
        if r.get("Entity") not in ("Keyword", "Product Targeting"):
            continue
        spend = _safe_float(r.get("Spend"))
        sales = _safe_float(r.get("Sales"))
        clicks = int(_safe_float(r.get("Clicks")))
        imps = int(_safe_float(r.get("Impressions")))
        orders = int(_safe_float(r.get("Orders")))
        cid = r.get("Campaign ID", "")
        camp_name = (r.get("Campaign Name") or
                     r.get("Campaign Name (Informational only)") or "")
        ag_name = (r.get("Ad Group Name") or
                   r.get("Ad Group Name (Informational only)") or "")
        targets.append({
            "entity": r.get("Entity", ""),
            "campaign_id": cid,
            "campaign_name": camp_name,
            "ad_group_id": r.get("Ad Group ID", ""),
            "ad_group_name": ag_name,
            "keyword_id": r.get("Keyword ID", ""),
            "product_targeting_id": r.get("Product Targeting ID", ""),
            "keyword_text": r.get("Keyword Text", ""),
            "match_type": r.get("Match Type", ""),
            "product_targeting_expression": r.get("Product Targeting Expression", ""),
            "bid": _safe_float(r.get("Bid")),
            "state": r.get("State", ""),
            "impressions": imps,
            "clicks": clicks,
            "spend": round(spend, 2),
            "sales": round(sales, 2),
            "orders": orders,
            "cvr": round(orders / clicks * 100, 2) if clicks else 0,
            "acos": round(spend / sales * 100, 2) if sales else 0,
            "cpc": round(spend / clicks, 2) if clicks else 0,
        })
    result["targets"] = targets

    # Placements: Bidding Adjustment rows
    placements = []
    for r in sp_data:
        if r.get("Entity") != "Bidding Adjustment":
            continue
        cid = r.get("Campaign ID", "")
        camp_name = (r.get("Campaign Name") or
                     r.get("Campaign Name (Informational only)") or "")
        spend = _safe_float(r.get("Spend"))
        sales = _safe_float(r.get("Sales"))
        clicks = int(_safe_float(r.get("Clicks")))
        imps = int(_safe_float(r.get("Impressions")))
        orders = int(_safe_float(r.get("Orders")))
        placements.append({
            "campaign_id": cid,
            "campaign_name": camp_name,
            "placement": r.get("Placement", ""),
            "percentage": _safe_float(r.get("Percentage")),
            "bidding_strategy": r.get("Bidding Strategy", ""),
            "impressions": imps,
            "clicks": clicks,
            "spend": round(spend, 2),
            "sales": round(sales, 2),
            "orders": orders,
            "cpc": round(spend / clicks, 2) if clicks else 0,
            "acos": round(spend / sales * 100, 2) if sales else 0,
            "cvr": round(orders / clicks * 100, 2) if clicks else 0,
        })
    result["placements"] = placements

    # Search terms with full per-campaign detail (for suggestions)
    search_terms_detail = []
    for r in st_data:
        term = r.get("Customer Search Term", "")
        if not term:
            continue
        spend = _safe_float(r.get("Spend"))
        sales = _safe_float(r.get("Sales"))
        clicks = int(_safe_float(r.get("Clicks")))
        imps = int(_safe_float(r.get("Impressions")))
        orders = int(_safe_float(r.get("Orders")))
        # Determine source match type
        match_type = (r.get("Match Type") or "").strip()
        pt_expr = (r.get("Product Targeting Expression") or "").strip()
        resolved = (r.get("Resolved Product Targeting Expression (Informational only)")
                    or "").strip()
        # Source type: if product targeting expression exists, it's auto/category
        if pt_expr and pt_expr.lower() not in ("", "none"):
            source_type = pt_expr  # e.g. "close-match", "loose-match", etc.
        elif match_type:
            source_type = match_type
        else:
            source_type = "unknown"
        camp_name = (r.get("Campaign Name (Informational only)") or "")
        ag_name = (r.get("Ad Group Name (Informational only)") or "")
        search_terms_detail.append({
            "search_term": term,
            "campaign_id": r.get("Campaign ID", ""),
            "campaign_name": camp_name,
            "ad_group_id": r.get("Ad Group ID", ""),
            "ad_group_name": ag_name,
            "source_type": source_type,
            "keyword_text": (r.get("Keyword Text") or ""),
            "impressions": imps,
            "clicks": clicks,
            "spend": round(spend, 2),
            "sales": round(sales, 2),
            "orders": orders,
            "cvr": round(orders / clicks * 100, 2) if clicks else 0,
            "acos": round(spend / sales * 100, 2) if sales else 0,
            "cpc": round(spend / clicks, 2) if clicks else 0,
        })
    result["search_terms_detail"] = search_terms_detail

    # Existing exact keywords (for dedup when creating exact campaigns)
    existing_exact = set()
    for r in sp_data:
        if r.get("Entity") == "Keyword":
            mt = (r.get("Match Type") or "").lower().strip()
            kw = (r.get("Keyword Text") or "").lower().strip()
            if mt == "exact" and kw:
                existing_exact.add(kw)
    result["existing_exact_keywords"] = list(existing_exact)

    # Existing negatives (for idempotency)
    existing_negatives = set()
    for r in sp_data:
        if r.get("Entity") in ("Campaign Negative Keyword", "Negative Keyword"):
            kw = (r.get("Keyword Text") or "").lower().strip()
            cid = r.get("Campaign ID", "")
            if kw and cid:
                existing_negatives.add(f"{cid}|{kw}")
    result["existing_negatives"] = list(existing_negatives)

    # Product Ads: ASIN + SKU per campaign/ad group
    product_ads = []
    for r in sp_data:
        if r.get("Entity") != "Product Ad":
            continue
        cid = r.get("Campaign ID", "")
        agid = r.get("Ad Group ID", "")
        sku = r.get("SKU", "") or ""
        asin = r.get("ASIN", "") or ""
        camp_name = (r.get("Campaign Name") or
                     r.get("Campaign Name (Informational only)") or "")
        ag_name = (r.get("Ad Group Name") or
                   r.get("Ad Group Name (Informational only)") or "")
        product_ads.append({
            "campaign_id": cid,
            "ad_group_id": agid,
            "campaign_name": camp_name,
            "ad_group_name": ag_name,
            "sku": sku,
            "asin": asin,
        })
    result["product_ads"] = product_ads

    # Portfolios: extract from Portfolios sheet or infer from campaign data
    portfolios = []
    port_headers, port_rows = _read_sheet(wb, "Portfolios")
    if port_headers:
        for r in _rows_to_dicts(port_headers, port_rows):
            pid = r.get("Portfolio ID", "")
            name = r.get("Portfolio Name", "") or r.get("Name", "")
            if pid:
                portfolios.append({"portfolio_id": str(pid), "name": name})
    else:
        # Infer from campaign rows – collect unique portfolio IDs
        seen_pids = set()
        for r in sp_data:
            if r.get("Entity") == "Campaign":
                pid = r.get("Portfolio ID", "")
                if pid and str(pid) not in seen_pids:
                    seen_pids.add(str(pid))
                    portfolios.append({"portfolio_id": str(pid), "name": ""})
    result["portfolios"] = portfolios

    wb.close()
    return result

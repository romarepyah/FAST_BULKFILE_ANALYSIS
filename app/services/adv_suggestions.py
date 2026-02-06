"""
Advanced Suggestions Engine for Fast Bulk Analysis.

Analyses data extracted from an Amazon bulk XLSX file (in-memory, no DB)
and returns actionable suggestions grouped into categories A-F.
Each suggestion carries bulk-file-ready action rows.
"""

import logging
import re
from collections import defaultdict
from datetime import date
from typing import Dict, List

logger = logging.getLogger(__name__)

# ── Default thresholds (user can override from frontend) ──────
DEFAULTS = {
    "acos_ineffective": 0.35,      # 35% – placement is "ineffective"
    "acos_target": 0.30,           # 30% – target ACOS for optimization
    "spend_campaign_pause": 15,    # $ – min spend to consider pausing campaign
    "spend_target_pause": 10,      # $ – min spend to consider pausing target
    "clicks_negative": 10,         # min clicks for negative suggestion
    "spend_negative": 5,           # $ – min spend for negative suggestion
    "negative_match_type": "Negative Exact",
    "ctr_negative": 0.002,         # 0.2% – CTR threshold
    "orders_create_exact": 2,      # min orders for "create exact campaign"
    "cvr_create_exact": 0.20,      # 20% – min CVR for creating exact
    "bid_multiplier": 1.1,         # multiplier on CPC for new campaign bids
    "cvr_bid_increase": 0.30,      # 30% – CVR for bid increase
    "acos_bid_increase": 0.20,     # 20% – max ACOS for bid increase
    "acos_target_increase": 0.25,  # 25% – ACOS ceiling when boosting bids
    "bid_increase_step": 0.15,     # 15% step for bid increase
    "orders_bid_increase": 3,      # min orders for bid increase
    "clicks_bid_increase": 10,     # min clicks for bid increase
    "max_placement_pct": 900,      # Amazon max placement percentage
    "bid_reduction_ratio": 0.5,    # how much to reduce bids in placement optimization
    "acos_pause": 1.0,             # max ACOS (fraction) before pause suggestion
    "custom_rules": None,          # user-defined rules from frontend
}


# ══════════════════════════════════════════════════════════════
# CUSTOM RULES EVALUATION
# ══════════════════════════════════════════════════════════════

def _evaluate_condition(data: dict, condition: dict) -> bool:
    """Evaluate a single condition against a data row."""
    metric = condition.get("metric", "")
    operator = condition.get("operator", ">")
    value = condition.get("value", 0)

    # Get the actual value from data (handle % metrics)
    actual = data.get(metric, 0)
    if actual is None:
        actual = 0

    # Convert percentage metrics (acos, cvr, ctr are stored as % in data)
    # but user enters them as whole numbers (e.g. 35 for 35%)
    # The data already has them as percentages, so compare directly

    try:
        actual = float(actual)
        value = float(value)
    except (TypeError, ValueError):
        return False

    if operator == ">":
        return actual > value
    elif operator == ">=":
        return actual >= value
    elif operator == "<":
        return actual < value
    elif operator == "<=":
        return actual <= value
    elif operator == "==":
        return actual == value
    elif operator == "!=":
        return actual != value
    return False


def _evaluate_rule(data: dict, rule: dict) -> bool:
    """Evaluate if all conditions in a rule are met."""
    if not rule.get("enabled", True):
        return False
    conditions = rule.get("conditions", [])
    if not conditions:
        return False
    return all(_evaluate_condition(data, c) for c in conditions)


def _find_matching_rule(data: dict, rules: list) -> dict | None:
    """Find the first matching enabled rule for a data row."""
    for rule in rules:
        if _evaluate_rule(data, rule):
            return rule
    return None


def _sf(v, d=0.0):
    """Safe float."""
    try:
        return float(v)
    except (TypeError, ValueError):
        return d


def _norm_kw(text: str) -> str:
    """Normalise keyword for comparison: lowercase, collapse whitespace."""
    return re.sub(r"\s+", " ", (text or "").strip().lower())


def _extract_asin(text: str) -> str:
    """Try to extract an Amazon ASIN (B0...) from a string."""
    m = re.search(r'\b(B0[A-Z0-9]{8})\b', text or "")
    return m.group(1) if m else ""


# ── Bulk row helpers ──────────────────────────────────────────

def _sp_row(**kw) -> dict:
    """Build a Sponsored Products bulk row dict."""
    row = {"Product": "Sponsored Products"}
    row.update(kw)
    return row


# ══════════════════════════════════════════════════════════════
# PUBLIC API
# ══════════════════════════════════════════════════════════════

def generate_suggestions(analysis: dict, thresholds: dict | None = None) -> List[Dict]:
    """
    Given the output of fast_analysis.analyze_bulk_file(), produce suggestions.
    *thresholds* can override any key from DEFAULTS.
    """
    t = {**DEFAULTS, **(thresholds or {})}
    suggestions: List[Dict] = []

    campaigns = analysis.get("campaigns_table", [])
    targets = analysis.get("targets", [])
    placements = analysis.get("placements", [])
    search_terms = analysis.get("search_terms_detail", [])
    existing_exact = set(_norm_kw(k) for k in analysis.get("existing_exact_keywords", []))
    existing_negatives = set(analysis.get("existing_negatives", []))

    # Build campaign lookup for cross-referencing
    camp_by_id = {c["campaign_id"]: c for c in campaigns}

    # Build campaign → ASIN/SKU lookup from product ads
    product_ads = analysis.get("product_ads", [])
    camp_asin: Dict[str, str] = {}
    camp_sku: Dict[str, str] = {}
    camp_all_skus: Dict[str, list] = defaultdict(list)
    for pa in product_ads:
        cid = pa["campaign_id"]
        if pa.get("asin") and cid not in camp_asin:
            camp_asin[cid] = pa["asin"]
        if pa.get("sku") and cid not in camp_sku:
            camp_sku[cid] = pa["sku"]
        if pa.get("sku") and pa["sku"] not in camp_all_skus[cid]:
            camp_all_skus[cid].append(pa["sku"])

    # Build portfolio lookup
    portfolios = analysis.get("portfolios", [])
    portfolio_by_id = {str(p["portfolio_id"]): p["name"] for p in portfolios}

    # Build campaign → portfolio_id lookup
    camp_portfolio: Dict[str, str] = {}
    for c in campaigns:
        pid = c.get("portfolio_id", "")
        if pid:
            camp_portfolio[str(c["campaign_id"])] = str(pid)

    # Count targets per campaign for "single target → pause campaign" logic
    targets_per_campaign = defaultdict(int)
    for tgt in targets:
        targets_per_campaign[tgt["campaign_id"]] += 1

    # Group targets by campaign (for bid update rows on Keyword/Product Targeting)
    targets_by_camp: Dict[str, list] = defaultdict(list)
    for tgt in targets:
        cid = str(tgt["campaign_id"])
        # Only include enabled targets with a bid
        if tgt.get("state", "").lower() == "enabled" and tgt.get("bid", 0) > 0:
            targets_by_camp[cid].append(tgt)

    # A) Create exact campaigns from search terms
    suggestions += _create_exact_campaigns(
        search_terms, existing_exact, camp_by_id, camp_asin, camp_sku,
        camp_all_skus, camp_portfolio, portfolio_by_id, t,
    )

    # B) Search term negatives
    suggestions += _negative_search_terms(
        search_terms, existing_negatives, t,
    )

    # C) Pause campaigns
    suggestions += _pause_campaigns(campaigns, t)

    # D) Pause targets
    suggestions += _pause_targets(targets, targets_per_campaign, t)

    # E) Placement optimization
    suggestions += _optimize_placements(campaigns, placements, targets_by_camp, t)

    # F) Bid increase for strong campaigns
    suggestions += _increase_bids(campaigns, placements, targets_by_camp, t)

    # Assign stable IDs
    for i, s in enumerate(suggestions):
        s["id"] = f"adv_{i}_{hash(s['title']) % 100000}"

    logger.info("Generated %d advanced suggestions", len(suggestions))
    return suggestions


# ══════════════════════════════════════════════════════════════
# A) CREATE EXACT CAMPAIGNS FROM SEARCH TERMS
# ══════════════════════════════════════════════════════════════

def _is_non_exact_source(source_type: str) -> bool:
    """Return True if the source is broad/phrase/auto/category (not exact)."""
    s = (source_type or "").lower().strip()
    # Exact sources
    if s in ("exact", "exact match"):
        return False
    # Non-exact sources: broad, phrase, auto types, category targets
    return True


def _create_exact_campaigns(
    search_terms: list, existing_exact: set,
    camp_by_id: dict, camp_asin: dict, camp_sku: dict,
    camp_all_skus: dict, camp_portfolio: dict, portfolio_by_id: dict,
    t: dict,
) -> list:
    out = []
    start_date = date.today().strftime("%Y%m%d")
    custom_rules = (t.get("custom_rules") or {}).get("exact", [])

    for r in search_terms:
        source = r.get("source_type", "")
        if not _is_non_exact_source(source):
            continue
        orders = r["orders"]
        clicks = r["clicks"]
        spend = r["spend"]
        sales = r["sales"]
        cvr = orders / clicks if clicks else 0

        # Build data dict for rule evaluation (use percentages for cvr/acos)
        data = {
            "orders": orders,
            "clicks": clicks,
            "spend": spend,
            "sales": sales,
            "cvr": round(cvr * 100, 2),  # as percentage
            "acos": round(spend / sales * 100, 2) if sales > 0 else 0,
            "ctr": r.get("ctr", 0),
            "cpc": r.get("cpc", 0),
            "impressions": r.get("impressions", 0),
        }

        # Check custom rules first, then fall back to defaults
        matched_rule = _find_matching_rule(data, custom_rules)
        bid_multiplier = t.get("bid_multiplier", 1.1)

        if matched_rule:
            # Use rule's action parameters
            bid_multiplier = matched_rule.get("action", {}).get("bid_multiplier", bid_multiplier)
        else:
            # Default behavior: check thresholds
            if orders <= t["orders_create_exact"]:
                continue
            if cvr < t["cvr_create_exact"]:
                continue

        term_norm = _norm_kw(r["search_term"])
        if not term_norm:
            continue
        if term_norm in existing_exact:
            continue
        # Mark as seen to prevent duplicates within the same run
        existing_exact.add(term_norm)

        acos = round(spend / sales * 100, 2) if sales else 0
        cpc = round(spend / clicks, 2) if clicks else 0.50
        suggested_bid = round(cpc * bid_multiplier, 2)
        camp = camp_by_id.get(r["campaign_id"], {})
        camp_budget = camp.get("daily_budget", 10)
        new_budget = round(max(5, camp_budget * 0.5), 2)

        # Get ASIN and SKU from source campaign's product ads
        source_cid = str(r["campaign_id"])
        asin = camp_asin.get(source_cid, "")
        if not asin:
            # Fallback: extract ASIN from source campaign name
            asin = _extract_asin(r.get("campaign_name", ""))
        if not asin:
            # Fallback: extract from any campaign name in the data
            asin = _extract_asin(camp.get("name", ""))
        sku = camp_sku.get(source_cid, "")
        available_skus = camp_all_skus.get(source_cid, [])

        # Portfolio from source campaign
        source_portfolio_id = camp_portfolio.get(source_cid, "")
        source_portfolio_name = portfolio_by_id.get(source_portfolio_id, "")

        # Campaign naming: SP Kw Ex {keyword} - {ASIN}
        kw_display = r["search_term"][:60]
        if asin:
            new_camp_name = f"SP Kw Ex {kw_display} - {asin}"
        else:
            new_camp_name = f"SP Kw Ex {kw_display}"
        new_ag_name = new_camp_name

        # For Create operations, Amazon links rows by
        # Campaign ID = Campaign Name, Ad Group ID = Ad Group Name
        new_camp_id = new_camp_name
        new_ag_id = new_ag_name

        actions = [
            # 1. Create campaign
            _sp_row(
                Entity="Campaign", Operation="Create",
                **{"Campaign ID": new_camp_id},
                **{"Campaign Name": new_camp_name},
                **{"Portfolio ID": source_portfolio_id},
                **{"Start Date": start_date},
                **{"Targeting Type": "Manual"},
                State="enabled",
                **{"Daily Budget": new_budget},
                **{"Bidding Strategy": "Dynamic bids - down only"},
            ),
            # 2. Create ad group
            _sp_row(
                Entity="Ad Group", Operation="Create",
                **{"Campaign ID": new_camp_id},
                **{"Ad Group ID": new_ag_id},
                **{"Campaign Name": new_camp_name},
                **{"Ad Group Name": new_ag_name},
                State="enabled",
                **{"Ad Group Default Bid": suggested_bid},
            ),
            # 3. Create product ad (with SKU/ASIN from source)
            _sp_row(
                Entity="Product Ad", Operation="Create",
                **{"Campaign ID": new_camp_id},
                **{"Ad Group ID": new_ag_id},
                **{"Campaign Name": new_camp_name},
                **{"Ad Group Name": new_ag_name},
                SKU=sku,
                ASIN=asin,
                State="enabled",
            ),
            # 4. Create exact keyword
            _sp_row(
                Entity="Keyword", Operation="Create",
                **{"Campaign ID": new_camp_id},
                **{"Ad Group ID": new_ag_id},
                **{"Campaign Name": new_camp_name},
                **{"Ad Group Name": new_ag_name},
                **{"Keyword Text": r["search_term"]},
                **{"Match Type": "Exact"},
                State="enabled",
                Bid=suggested_bid,
            ),
            # 5. Add negative exact in SOURCE campaign (not the new one)
            _sp_row(
                Entity="Campaign Negative Keyword", Operation="Create",
                **{"Campaign ID": r["campaign_id"]},
                **{"Campaign Name": r["campaign_name"]},
                **{"Keyword Text": r["search_term"]},
                **{"Match Type": "Negative Exact"},
                State="enabled",
            ),
        ]

        out.append({
            "category": "Create Exact Campaign",
            "severity": "medium",
            "title": (f"Create exact campaign for '{r['search_term'][:60]}' — "
                      f"{orders} orders, CVR {cvr*100:.0f}%, ACOS {acos:.0f}%"),
            "detail": (f"This search term from '{r['campaign_name']}' converts well via "
                       f"{source}. Create a dedicated exact campaign and negative it in the source."),
            "metrics": {
                "search_term": r["search_term"],
                "orders": orders, "clicks": clicks,
                "cvr": round(cvr * 100, 1), "acos": acos,
                "spend": spend, "sales": sales,
                "suggested_bid": suggested_bid,
                "asin": asin,
                "sku": sku,
                "available_skus": available_skus,
                "source_portfolio_id": source_portfolio_id,
                "source_portfolio_name": source_portfolio_name,
                "amazon_url": f"https://www.amazon.com/s?k={r['search_term'].replace(' ', '+')}",
            },
            "actions": actions,
        })
    return out


# ══════════════════════════════════════════════════════════════
# B) SEARCH TERM NEGATIVES
# ══════════════════════════════════════════════════════════════

def _negative_search_terms(search_terms: list, existing_negatives: set, t: dict) -> list:
    out = []
    seen = set()
    custom_rules = (t.get("custom_rules") or {}).get("negatives", [])

    for r in search_terms:
        source = r.get("source_type", "")
        if not _is_non_exact_source(source):
            continue

        clicks = r["clicks"]
        orders = r["orders"]
        spend = r["spend"]
        sales = r.get("sales", 0)

        # Build data dict for rule evaluation
        data = {
            "clicks": clicks,
            "orders": orders,
            "spend": spend,
            "sales": sales,
            "acos": round(spend / sales * 100, 2) if sales > 0 else 0,
            "cvr": round(orders / clicks * 100, 2) if clicks > 0 else 0,
            "ctr": r.get("ctr", 0),
            "cpc": r.get("cpc", 0),
            "impressions": r.get("impressions", 0),
        }

        # Check custom rules first
        matched_rule = _find_matching_rule(data, custom_rules)
        if matched_rule:
            match_type = matched_rule.get("action", {}).get("match_type", "Negative Exact")
        elif clicks >= t["clicks_negative"] and orders == 0:
            # Default behavior
            match_type = t.get("negative_match_type", "Negative Exact")
        else:
            continue

        term = r["search_term"]
        cid = str(r["campaign_id"])
        dedup_key = f"{cid}|{_norm_kw(term)}"
        if dedup_key in existing_negatives or dedup_key in seen:
            continue
        seen.add(dedup_key)

        out.append({
            "category": "Search Term Negatives",
            "severity": "high",
            "title": (f"Negative '{term[:60]}' in '{r['campaign_name']}' — "
                      f"{clicks} clicks, ${spend:.2f}, 0 orders"),
            "detail": (f"Wasting spend via {source}. Add as campaign-level negative exact."),
            "metrics": {
                "search_term": term, "clicks": clicks,
                "spend": spend, "source": source,
                "cpc": r.get("cpc", 0),
                "amazon_url": f"https://www.amazon.com/s?k={term.replace(' ', '+')}",
            },
            "actions": [
                _sp_row(
                    Entity="Campaign Negative Keyword", Operation="Create",
                    **{"Campaign ID": r["campaign_id"]},
                    **{"Campaign Name": r["campaign_name"]},
                    **{"Keyword Text": term},
                    **{"Match Type": match_type},
                    State="enabled",
                ),
            ],
        })
    return out


# ══════════════════════════════════════════════════════════════
# C) PAUSE CAMPAIGNS
# ══════════════════════════════════════════════════════════════

def _pause_campaigns(campaigns: list, t: dict) -> list:
    out = []
    custom_rules = (t.get("custom_rules") or {}).get("pause", [])

    for c in campaigns:
        spend = c["spend"]
        orders = c["orders"]
        clicks = c["clicks"]
        sales = c["sales"]

        # Build data dict for rule evaluation
        data = {
            "spend": spend,
            "orders": orders,
            "clicks": clicks,
            "sales": sales,
            "impressions": c["impressions"],
            "acos": round(spend / sales * 100, 2) if sales > 0 else 0,
            "cvr": round(orders / clicks * 100, 2) if clicks > 0 else 0,
            "ctr": c.get("ctr", 0),
            "cpc": c.get("cpc", 0),
            "roas": c.get("roas", 0),
        }

        # Check custom rules first, then fall back to defaults
        matched_rule = _find_matching_rule(data, custom_rules)
        if not matched_rule:
            # Default behavior: pause if spend >= threshold and 0 orders
            if spend < t["spend_campaign_pause"] or orders > 0:
                continue

        out.append({
            "category": "Pause Campaigns",
            "severity": "high",
            "title": f"Pause '{c['name'][:50]}' — ${spend:.2f} spend, 0 orders",
            "detail": "Campaign is spending with no conversions.",
            "metrics": {
                "spend": spend, "orders": orders,
                "clicks": clicks, "impressions": c["impressions"],
                "cpc": c.get("cpc", 0),
            },
            "actions": [
                _sp_row(
                    Entity="Campaign", Operation="Update",
                    **{"Campaign ID": c["campaign_id"]},
                    **{"Campaign Name": c["name"]},
                    State="paused",
                ),
            ],
        })
    return out


# ══════════════════════════════════════════════════════════════
# D) PAUSE TARGETS
# ══════════════════════════════════════════════════════════════

def _pause_targets(targets: list, targets_per_campaign: dict, t: dict) -> list:
    out = []
    custom_rules = (t.get("custom_rules") or {}).get("pause", [])

    for tgt in targets:
        spend = tgt["spend"]
        orders = tgt["orders"]
        clicks = tgt["clicks"]
        sales = tgt.get("sales", 0)

        # Build data dict for rule evaluation
        data = {
            "spend": spend,
            "orders": orders,
            "clicks": clicks,
            "sales": sales,
            "impressions": tgt.get("impressions", 0),
            "acos": round(spend / sales * 100, 2) if sales > 0 else 0,
            "cvr": round(orders / clicks * 100, 2) if clicks > 0 else 0,
            "cpc": tgt.get("cpc", 0),
            "bid": tgt.get("bid", 0),
        }

        # Check custom rules first, then fall back to defaults
        matched_rule = _find_matching_rule(data, custom_rules)
        if not matched_rule:
            # Default behavior: pause if spend >= threshold and 0 orders
            if spend < t["spend_target_pause"] or orders > 0:
                continue

        cid = tgt["campaign_id"]
        is_sole_target = targets_per_campaign.get(cid, 0) <= 1
        target_label = tgt["keyword_text"] or tgt["product_targeting_expression"] or "?"

        if is_sole_target:
            # Single target in campaign → pause the campaign instead
            out.append({
                "category": "Pause Targets",
                "severity": "high",
                "title": (f"Pause campaign '{tgt['campaign_name'][:40]}' "
                          f"(sole target '{target_label[:30]}') — ${spend:.2f}, 0 orders"),
                "detail": "Only target in campaign is unprofitable. Pause the entire campaign.",
                "metrics": {
                    "target": target_label, "spend": spend,
                    "clicks": tgt["clicks"], "orders": 0,
                },
                "actions": [
                    _sp_row(
                        Entity="Campaign", Operation="Update",
                        **{"Campaign ID": cid},
                        **{"Campaign Name": tgt["campaign_name"]},
                        State="paused",
                    ),
                ],
            })
        else:
            entity = tgt["entity"]  # "Keyword" or "Product Targeting"
            action = _sp_row(
                Entity=entity, Operation="Update",
                **{"Campaign ID": cid},
                **{"Ad Group ID": tgt["ad_group_id"]},
                State="paused",
            )
            if entity == "Keyword":
                action["Keyword Text"] = tgt["keyword_text"]
                action["Match Type"] = tgt["match_type"]
            else:
                action["Product Targeting Expression"] = tgt["product_targeting_expression"]

            out.append({
                "category": "Pause Targets",
                "severity": "high",
                "title": (f"Pause {entity.lower()} '{target_label[:40]}' in "
                          f"'{tgt['campaign_name'][:30]}' — ${spend:.2f}, 0 orders"),
                "detail": f"This {entity.lower()} is wasting spend with no conversions.",
                "metrics": {
                    "target": target_label, "spend": spend,
                    "clicks": tgt["clicks"], "orders": 0,
                    "campaign": tgt["campaign_name"],
                },
                "actions": [action],
            })
    return out


# ══════════════════════════════════════════════════════════════
# E) PLACEMENT OPTIMIZATION
# ══════════════════════════════════════════════════════════════

def _optimize_placements(
    campaigns: list, placements: list, targets_by_camp: dict, t: dict,
) -> list:
    """
    For campaigns with placement breakdown, find ineffective placements
    (ACOS > threshold) and suggest bid/multiplier adjustments.
    Includes Keyword/Product Targeting bid update rows.
    """
    out = []
    custom_rules = (t.get("custom_rules") or {}).get("placement", [])

    # Group placements by campaign
    plc_by_camp = defaultdict(list)
    for p in placements:
        plc_by_camp[p["campaign_id"]].append(p)

    camp_by_id = {c["campaign_id"]: c for c in campaigns}

    for cid, camp_placements in plc_by_camp.items():
        camp = camp_by_id.get(cid)
        if not camp:
            continue
        if camp["spend"] < 5:
            continue

        # Classify placements using custom rules or default thresholds
        effective = []
        ineffective = []
        for p in camp_placements:
            if p["spend"] <= 0:
                continue

            # Build data dict for rule evaluation
            data = {
                "acos": p["acos"] if p["acos"] else 0,
                "spend": p["spend"],
                "sales": p.get("sales", 0),
                "clicks": p.get("clicks", 0),
                "orders": p.get("orders", 0),
                "impressions": p.get("impressions", 0),
                "percentage": int(p.get("percentage", 0)),
                "cvr": p.get("cvr", 0),
                "ctr": p.get("ctr", 0),
                "cpc": p.get("cpc", 0),
            }

            # Check custom rules first
            matched_rule = _find_matching_rule(data, custom_rules)
            if matched_rule:
                # Rule matched - consider ineffective if percentage > 0
                if int(p.get("percentage", 0)) > 0:
                    ineffective.append((p, matched_rule))
            else:
                # Default behavior
                acos_frac = p["acos"] / 100 if p["acos"] else 0
                if acos_frac > t["acos_ineffective"]:
                    if int(p.get("percentage", 0)) > 0:
                        ineffective.append((p, None))
                else:
                    effective.append(p)

        if not ineffective:
            continue

        # Build lookup for ineffective placements and their rules
        ineffective_map = {p["placement"]: (p, rule) for p, rule in ineffective}

        # Find best placement (lowest ACOS with meaningful spend)
        best = None
        if effective:
            best = min(effective, key=lambda x: x["acos"] if x["acos"] > 0 else 9999)

        # Compute recommended adjustments
        camp_cpc = camp.get("cpc", 0)
        if camp_cpc <= 0:
            continue

        best_pct = best["percentage"] if best else 0
        best_factor = 1 + (best_pct / 100)
        base_bid = round(camp_cpc / best_factor, 2) if best_factor > 0 else camp_cpc

        # New strategy: halve the base bid, raise best placement to maintain CPC
        bid_reduction = t.get("bid_reduction_ratio", 0.5)
        new_base = round(base_bid * bid_reduction, 2)
        if new_base < 0.02:
            new_base = 0.02

        old_effective = base_bid * best_factor
        desired_factor = old_effective / new_base if new_base > 0 else 1
        new_best_pct = round((desired_factor - 1) * 100)
        new_best_pct = min(new_best_pct, t["max_placement_pct"])

        actions = []
        placement_changes = []

        # 1) Keyword / Product Targeting bid update rows
        bid_ratio = new_base / base_bid if base_bid > 0 else 1
        for tgt in targets_by_camp.get(str(cid), []):
            old_bid = tgt.get("bid", 0)
            new_bid = round(old_bid * bid_ratio, 2)
            if new_bid < 0.02:
                new_bid = 0.02
            action = _sp_row(
                Entity=tgt["entity"], Operation="Update",
                **{"Campaign ID": cid},
                **{"Ad Group ID": tgt["ad_group_id"]},
                **{"Campaign Name": camp["name"]},
                **{"Ad Group Name": tgt.get("ad_group_name", "")},
                Bid=new_bid,
                State="enabled",
            )
            if tgt["entity"] == "Keyword":
                action["Keyword ID"] = tgt.get("keyword_id", "")
                action["Keyword Text"] = tgt["keyword_text"]
                action["Match Type"] = tgt["match_type"]
            else:
                action["Product Targeting ID"] = tgt.get("product_targeting_id", "")
                action["Product Targeting Expression"] = tgt["product_targeting_expression"]
            actions.append(action)

        # 2) Placement adjustment rows
        for p in camp_placements:
            placement_name = p["placement"]
            old_pct = int(p["percentage"])

            # Check if this placement is in our ineffective map
            if placement_name in ineffective_map:
                ineff_p, matched_rule = ineffective_map[placement_name]
                if matched_rule:
                    # Use rule's action to determine new percentage
                    action_type = matched_rule.get("action", {}).get("type", "set_percentage")
                    if action_type == "set_percentage":
                        new_pct = int(matched_rule.get("action", {}).get("value", 0))
                    elif action_type == "reduce_bid":
                        # reduce_bid action applies to bids, not percentage
                        new_pct = 0  # still reduce placement
                    else:
                        new_pct = 0
                else:
                    # Default: set to 0
                    new_pct = 0
            elif best and p["placement"] == best["placement"]:
                new_pct = new_best_pct
            else:
                new_pct = old_pct

            if new_pct != old_pct:
                placement_changes.append(
                    f"{placement_name}: {old_pct}% → {new_pct}%"
                )
                actions.append(_sp_row(
                    Entity="Bidding Adjustment", Operation="Update",
                    **{"Campaign ID": cid},
                    **{"Campaign Name": camp["name"]},
                    Placement=placement_name,
                    Percentage=new_pct,
                ))

        if not actions:
            continue

        ineff_names = ", ".join(p["placement"] for p, _ in ineffective)
        out.append({
            "category": "Placement Optimization",
            "severity": "medium",
            "title": (f"Adjust placements for '{camp['name'][:40]}' — "
                      f"ineffective: {ineff_names}"),
            "detail": (f"Reduce exposure on high-ACOS placements. "
                       f"Changes: {'; '.join(placement_changes)}. "
                       f"Base bid ${base_bid:.2f} → ${new_base:.2f}."),
            "metrics": {
                "campaign_spend": camp["spend"],
                "campaign_acos": camp.get("acos", 0),
                "ineffective_placements": len(ineffective),
                "new_base_bid": new_base,
            },
            "actions": actions,
        })

    return out


# ══════════════════════════════════════════════════════════════
# F) INCREASE BIDS FOR STRONG CAMPAIGNS
# ══════════════════════════════════════════════════════════════

def _increase_bids(
    campaigns: list, placements: list, targets_by_camp: dict, t: dict,
) -> list:
    out = []
    custom_rules = (t.get("custom_rules") or {}).get("bids", [])

    plc_by_camp = defaultdict(list)
    for p in placements:
        plc_by_camp[p["campaign_id"]].append(p)

    min_clicks = t.get("clicks_bid_increase", 10)
    min_orders = t.get("orders_bid_increase", 3)

    for c in campaigns:
        clicks = c["clicks"]
        orders = c["orders"]
        spend = c["spend"]
        sales = c["sales"]
        if sales <= 0:
            continue

        cvr = orders / clicks if clicks > 0 else 0
        acos_frac = spend / sales if sales > 0 else 1

        # Build data dict for rule evaluation
        data = {
            "clicks": clicks,
            "orders": orders,
            "spend": spend,
            "sales": sales,
            "cvr": round(cvr * 100, 2),  # as percentage
            "acos": round(acos_frac * 100, 2),  # as percentage
            "ctr": c.get("ctr", 0),
            "cpc": c.get("cpc", 0),
            "impressions": c.get("impressions", 0),
            "roas": c.get("roas", 0),
        }

        # Check custom rules first
        matched_rule = _find_matching_rule(data, custom_rules)
        step = t["bid_increase_step"]

        if matched_rule:
            # Use rule's action step
            rule_step = matched_rule.get("action", {}).get("step", 15)
            step = rule_step / 100  # convert from percentage
        else:
            # Default behavior: check thresholds
            if clicks < min_clicks or orders < min_orders:
                continue
            if cvr < t["cvr_bid_increase"] or acos_frac > t["acos_bid_increase"]:
                continue

        cpc = c.get("cpc", 0)
        if cpc <= 0:
            continue

        # Calculate how much we can increase CPC and stay under target ACOS
        aov = sales / orders if orders else 0
        max_cpc = t["acos_target_increase"] * cvr * aov
        if max_cpc <= cpc:
            continue  # already at or above the ceiling

        new_cpc = round(min(cpc * (1 + step), max_cpc), 2)
        bid_ratio = new_cpc / cpc if cpc > 0 else 1

        cid = c["campaign_id"]
        actions = []

        # 1) Keyword / Product Targeting bid update rows
        for tgt in targets_by_camp.get(str(cid), []):
            old_bid = tgt.get("bid", 0)
            new_bid = round(old_bid * bid_ratio, 2)
            action = _sp_row(
                Entity=tgt["entity"], Operation="Update",
                **{"Campaign ID": cid},
                **{"Ad Group ID": tgt["ad_group_id"]},
                **{"Campaign Name": c["name"]},
                **{"Ad Group Name": tgt.get("ad_group_name", "")},
                Bid=new_bid,
                State="enabled",
            )
            if tgt["entity"] == "Keyword":
                action["Keyword ID"] = tgt.get("keyword_id", "")
                action["Keyword Text"] = tgt["keyword_text"]
                action["Match Type"] = tgt["match_type"]
            else:
                action["Product Targeting ID"] = tgt.get("product_targeting_id", "")
                action["Product Targeting Expression"] = tgt["product_targeting_expression"]
            actions.append(action)

        # 2) Placement adjustment rows (boost best placement proportionally)
        camp_placements = plc_by_camp.get(cid, [])
        if camp_placements:
            best = min(
                [p for p in camp_placements if p["spend"] > 0],
                key=lambda x: x["acos"] if x["acos"] > 0 else 9999,
                default=None,
            )
            if best:
                old_factor = 1 + (best["percentage"] / 100)
                new_factor = new_cpc / cpc * old_factor
                new_pct = round((new_factor - 1) * 100)
                new_pct = min(new_pct, t["max_placement_pct"])
                if new_pct != int(best["percentage"]):
                    actions.append(_sp_row(
                        Entity="Bidding Adjustment", Operation="Update",
                        **{"Campaign ID": cid},
                        **{"Campaign Name": c["name"]},
                        Placement=best["placement"],
                        Percentage=new_pct,
                    ))

        if not actions:
            continue

        out.append({
            "category": "Increase Bids",
            "severity": "low",
            "title": (f"Boost '{c['name'][:40]}' — CVR {cvr*100:.0f}%, "
                      f"ACOS {acos_frac*100:.0f}%, {orders} orders"),
            "detail": (f"Strong campaign. Suggested CPC increase: "
                       f"${cpc:.2f} → ${new_cpc:.2f}. "
                       f"Max CPC at {t['acos_target_increase']*100:.0f}% "
                       f"ACOS target: ${max_cpc:.2f}."),
            "metrics": {
                "cvr": round(cvr * 100, 1),
                "acos": round(acos_frac * 100, 1),
                "orders": orders, "spend": spend, "sales": sales,
                "current_cpc": cpc, "suggested_cpc": new_cpc,
                "max_cpc": round(max_cpc, 2),
            },
            "actions": actions,
        })

    return out


# ══════════════════════════════════════════════════════════════
# BULK CSV GENERATION
# ══════════════════════════════════════════════════════════════

# Column order matching Amazon's bulk file format
BULK_COLS = [
    "Product", "Entity", "Operation",
    "Campaign ID", "Ad Group ID", "Portfolio ID",
    "Ad ID", "Keyword ID", "Product Targeting ID",
    "Campaign Name", "Ad Group Name",
    "Start Date", "End Date",
    "Targeting Type", "State",
    "Daily Budget", "SKU", "ASIN",
    "Ad Group Default Bid", "Bid",
    "Keyword Text", "Match Type",
    "Bidding Strategy", "Placement", "Percentage",
    "Product Targeting Expression",
]


def build_bulk_xlsx(selected_suggestions: List[Dict]) -> bytes:
    """
    Given a list of suggestion dicts (each with an "actions" list),
    produce an Excel (.xlsx) file as bytes ready for download.
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    rows = []
    for sug in selected_suggestions:
        for action in sug.get("actions", []):
            row = {}
            for col in BULK_COLS:
                val = action.get(col, "")
                if val is None:
                    val = ""
                row[col] = val
            if not row["Product"]:
                row["Product"] = "Sponsored Products"
            rows.append(row)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sponsored Products Campaigns"

    # Header row
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2",
                              fill_type="solid")
    for col_idx, col_name in enumerate(BULK_COLS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, col_name in enumerate(BULK_COLS, 1):
            val = row_data.get(col_name, "")
            # Keep numeric types as numbers for Excel
            if isinstance(val, (int, float)):
                ws.cell(row=row_idx, column=col_idx, value=val)
            else:
                ws.cell(row=row_idx, column=col_idx, value=str(val))

    # Auto-fit column widths (approximate)
    for col_idx, col_name in enumerate(BULK_COLS, 1):
        max_len = len(col_name)
        for row_idx in range(2, len(rows) + 2):
            cell_val = str(ws.cell(row=row_idx, column=col_idx).value or "")
            max_len = max(max_len, len(cell_val))
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()

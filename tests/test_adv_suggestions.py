"""Unit tests for the advanced suggestions engine (pure functions, no DB)."""

import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from app.services.adv_suggestions import (
    generate_suggestions,
    build_bulk_xlsx,
    _pause_campaigns,
    _pause_targets,
    _negative_search_terms,
    _create_exact_campaigns,
    _optimize_placements,
    _increase_bids,
    DEFAULTS,
)


# ── helpers ──────────────────────────────────────────────────

def _camp(name="C1", cid="100", spend=50, sales=100, orders=5,
          clicks=200, impressions=10000, daily_budget=10, cpc=0.25,
          state="enabled", targeting_type="Manual", **kw):
    c = {
        "campaign_id": cid, "name": name, "state": state,
        "targeting_type": targeting_type, "daily_budget": daily_budget,
        "impressions": impressions, "clicks": clicks,
        "spend": spend, "sales": sales, "orders": orders,
        "ctr": round(clicks/impressions*100, 2) if impressions else 0,
        "cpc": cpc,
        "acos": round(spend/sales*100, 2) if sales else 0,
        "roas": round(sales/spend, 2) if spend else 0,
    }
    c.update(kw)
    return c


def _target(entity="Keyword", cid="100", agid="200", spend=20, orders=0,
            clicks=30, keyword_text="test kw", match_type="Broad",
            product_targeting_expression="", bid=0.5, campaign_name="C1",
            ad_group_name="AG1", **kw):
    t = {
        "entity": entity, "campaign_id": cid, "ad_group_id": agid,
        "campaign_name": campaign_name, "ad_group_name": ad_group_name,
        "keyword_text": keyword_text, "match_type": match_type,
        "product_targeting_expression": product_targeting_expression,
        "bid": bid, "state": "enabled",
        "impressions": 1000, "clicks": clicks,
        "spend": spend, "sales": 0 if orders == 0 else spend * 2,
        "orders": orders,
        "cvr": round(orders/clicks*100, 2) if clicks else 0,
        "acos": 0,
        "cpc": round(spend/clicks, 2) if clicks else 0,
    }
    t.update(kw)
    return t


def _search_term(term="good term", cid="100", clicks=50, orders=5,
                 spend=10, sales=50, source_type="broad",
                 campaign_name="C1", ad_group_id="200", ad_group_name="AG1",
                 **kw):
    s = {
        "search_term": term, "campaign_id": cid, "ad_group_id": ad_group_id,
        "campaign_name": campaign_name, "ad_group_name": ad_group_name,
        "source_type": source_type, "keyword_text": "",
        "impressions": 1000, "clicks": clicks,
        "spend": spend, "sales": sales, "orders": orders,
        "cvr": round(orders/clicks*100, 2) if clicks else 0,
        "acos": round(spend/sales*100, 2) if sales else 0,
        "cpc": round(spend/clicks, 2) if clicks else 0,
    }
    s.update(kw)
    return s


def _placement(cid="100", placement="Placement Top", percentage=100,
               spend=50, sales=100, clicks=100, orders=5,
               campaign_name="C1", **kw):
    p = {
        "campaign_id": cid, "campaign_name": campaign_name,
        "placement": placement, "percentage": percentage,
        "bidding_strategy": "Fixed bid",
        "impressions": 5000, "clicks": clicks,
        "spend": spend, "sales": sales, "orders": orders,
        "cpc": round(spend/clicks, 2) if clicks else 0,
        "acos": round(spend/sales*100, 2) if sales else 0,
        "cvr": round(orders/clicks*100, 2) if clicks else 0,
    }
    p.update(kw)
    return p


# ══════════════════════════════════════════════════════════════
# TESTS
# ══════════════════════════════════════════════════════════════

class TestPauseCampaigns:
    def test_pauses_zero_order_campaign(self):
        camps = [_camp(spend=20, orders=0, sales=0)]
        result = _pause_campaigns(camps, DEFAULTS)
        assert len(result) == 1
        assert result[0]["category"] == "Pause Campaigns"
        assert result[0]["actions"][0]["State"] == "paused"

    def test_skips_profitable_campaign(self):
        camps = [_camp(spend=50, orders=10, sales=200)]
        result = _pause_campaigns(camps, DEFAULTS)
        assert len(result) == 0

    def test_skips_low_spend(self):
        camps = [_camp(spend=5, orders=0, sales=0)]
        result = _pause_campaigns(camps, DEFAULTS)
        assert len(result) == 0


class TestPauseTargets:
    def test_pauses_target_with_zero_orders(self):
        targets = [_target(spend=15, orders=0)]
        per_camp = {"100": 3}
        result = _pause_targets(targets, per_camp, DEFAULTS)
        assert len(result) == 1
        assert result[0]["actions"][0]["State"] == "paused"
        assert result[0]["actions"][0]["Entity"] == "Keyword"

    def test_pauses_campaign_when_sole_target(self):
        targets = [_target(spend=15, orders=0)]
        per_camp = {"100": 1}
        result = _pause_targets(targets, per_camp, DEFAULTS)
        assert len(result) == 1
        assert result[0]["actions"][0]["Entity"] == "Campaign"

    def test_skips_converting_target(self):
        targets = [_target(spend=15, orders=3)]
        per_camp = {"100": 3}
        result = _pause_targets(targets, per_camp, DEFAULTS)
        assert len(result) == 0


class TestNegativeSearchTerms:
    def test_negatives_wasting_term(self):
        terms = [_search_term(term="bad term", clicks=15, orders=0, spend=20, sales=0)]
        result = _negative_search_terms(terms, set(), DEFAULTS)
        assert len(result) == 1
        assert result[0]["actions"][0]["Entity"] == "Campaign Negative Keyword"
        assert result[0]["actions"][0]["Match Type"] == "Negative Exact"

    def test_skips_exact_source(self):
        terms = [_search_term(clicks=15, orders=0, spend=20, sales=0, source_type="exact")]
        result = _negative_search_terms(terms, set(), DEFAULTS)
        assert len(result) == 0

    def test_skips_converting_term(self):
        terms = [_search_term(clicks=15, orders=3, spend=20, sales=60)]
        result = _negative_search_terms(terms, set(), DEFAULTS)
        assert len(result) == 0

    def test_skips_existing_negative(self):
        terms = [_search_term(term="bad term", clicks=15, orders=0, spend=20, sales=0)]
        existing = {"100|bad term"}
        result = _negative_search_terms(terms, existing, DEFAULTS)
        assert len(result) == 0


class TestCreateExactCampaigns:
    def test_creates_exact_for_good_term(self):
        terms = [_search_term(
            term="great keyword", clicks=20, orders=5, spend=10, sales=100,
            source_type="broad",
        )]
        camp_by_id = {"100": _camp()}
        camp_asin = {"100": "B0TEST1234"}
        camp_sku = {"100": "SKU-TEST-01"}
        camp_all_skus = {"100": ["SKU-TEST-01", "SKU-TEST-02"]}
        result = _create_exact_campaigns(
            terms, set(), camp_by_id, camp_asin, camp_sku,
            camp_all_skus, {}, {}, DEFAULTS,
        )
        assert len(result) == 1
        actions = result[0]["actions"]
        # Should create: campaign, ad group, product ad, keyword, negative in source
        assert len(actions) == 5
        entities = [a["Entity"] for a in actions]
        assert "Campaign" in entities
        assert "Ad Group" in entities
        assert "Product Ad" in entities
        assert "Keyword" in entities
        assert "Campaign Negative Keyword" in entities
        # Verify naming includes ASIN
        camp_action = [a for a in actions if a["Entity"] == "Campaign"][0]
        assert "B0TEST1234" in camp_action["Campaign Name"]
        assert "great keyword" in camp_action["Campaign Name"]
        assert camp_action["Bidding Strategy"] == "Dynamic bids - down only"
        assert camp_action["Start Date"] != ""
        # For Create: Campaign ID = Campaign Name, Ad Group ID = Ad Group Name
        assert camp_action["Campaign ID"] == camp_action["Campaign Name"]
        ag_action = [a for a in actions if a["Entity"] == "Ad Group"][0]
        assert ag_action["Campaign ID"] == camp_action["Campaign Name"]
        assert ag_action["Ad Group ID"] == ag_action["Ad Group Name"]
        # Verify product ad has SKU and ASIN, and links to new campaign/ad group
        pa_action = [a for a in actions if a["Entity"] == "Product Ad"][0]
        assert pa_action["SKU"] == "SKU-TEST-01"
        assert pa_action["ASIN"] == "B0TEST1234"
        assert pa_action["Campaign ID"] == camp_action["Campaign Name"]
        assert pa_action["Ad Group ID"] == ag_action["Ad Group Name"]
        # Verify negative goes to SOURCE campaign, not the new one
        neg_action = [a for a in actions if a["Entity"] == "Campaign Negative Keyword"][0]
        assert neg_action["Campaign ID"] == "100"  # source campaign ID
        assert neg_action["Campaign Name"] == "C1"  # source campaign name
        # Verify metrics include available SKUs and amazon URL
        m = result[0]["metrics"]
        assert m["available_skus"] == ["SKU-TEST-01", "SKU-TEST-02"]
        assert "amazon.com" in m["amazon_url"]

    def test_extracts_asin_from_campaign_name(self):
        """When no product ads exist, ASIN should be extracted from campaign name."""
        terms = [_search_term(
            term="great keyword", clicks=20, orders=5, spend=10, sales=100,
            source_type="broad", campaign_name="SP Auto CM - B0TESTASIN",
        )]
        camp_by_id = {"100": _camp(name="SP Auto CM - B0TESTASIN")}
        result = _create_exact_campaigns(
            terms, set(), camp_by_id, {}, {}, {}, {}, {}, DEFAULTS,
        )
        assert len(result) == 1
        camp_action = [a for a in result[0]["actions"] if a["Entity"] == "Campaign"][0]
        assert "B0TESTASIN" in camp_action["Campaign Name"]

    def test_skips_existing_exact(self):
        terms = [_search_term(
            term="great keyword", clicks=20, orders=5, spend=10, sales=100,
        )]
        existing_exact = {"great keyword"}
        result = _create_exact_campaigns(
            terms, existing_exact, {}, {}, {}, {}, {}, {}, DEFAULTS,
        )
        assert len(result) == 0

    def test_skips_exact_source(self):
        terms = [_search_term(
            term="exact term", clicks=20, orders=5, spend=10, sales=100,
            source_type="exact",
        )]
        result = _create_exact_campaigns(
            terms, set(), {}, {}, {}, {}, {}, {}, DEFAULTS,
        )
        assert len(result) == 0

    def test_skips_low_cvr(self):
        terms = [_search_term(
            term="low cvr", clicks=100, orders=3, spend=50, sales=90,
            source_type="broad",
        )]
        result = _create_exact_campaigns(
            terms, set(), {}, {}, {}, {}, {}, {}, DEFAULTS,
        )
        assert len(result) == 0  # CVR = 3% < 20%

    def test_portfolio_propagation(self):
        """Portfolio ID from source campaign should be set on new campaign."""
        terms = [_search_term(
            term="great keyword", clicks=20, orders=5, spend=10, sales=100,
            source_type="broad",
        )]
        camp_by_id = {"100": _camp()}
        camp_portfolio = {"100": "P123"}
        portfolio_by_id = {"P123": "My Portfolio"}
        result = _create_exact_campaigns(
            terms, set(), camp_by_id, {}, {}, {}, camp_portfolio, portfolio_by_id, DEFAULTS,
        )
        assert len(result) == 1
        camp_action = [a for a in result[0]["actions"] if a["Entity"] == "Campaign"][0]
        assert camp_action["Portfolio ID"] == "P123"
        assert result[0]["metrics"]["source_portfolio_name"] == "My Portfolio"


class TestPlacementOptimization:
    def test_detects_ineffective_placement(self):
        camps = [_camp(cid="100", spend=50, cpc=1.0)]
        placements = [
            _placement(cid="100", placement="Placement Top", percentage=100,
                       spend=30, sales=120, clicks=30),  # ACOS 25% - good
            _placement(cid="100", placement="Placement Rest Of Search", percentage=50,
                       spend=20, sales=40, clicks=20),   # ACOS 50% - bad
        ]
        targets_by_camp = {"100": [
            _target(cid="100", keyword_text="kw1", match_type="Broad",
                    bid=0.80, spend=25, orders=3),
        ]}
        result = _optimize_placements(camps, placements, targets_by_camp, DEFAULTS)
        assert len(result) == 1
        assert "Placement Rest Of Search" in result[0]["title"]
        # Should include Keyword bid update rows
        kw_actions = [a for a in result[0]["actions"] if a["Entity"] == "Keyword"]
        assert len(kw_actions) >= 1
        assert "Bid" in kw_actions[0]
        assert "Keyword Text" in kw_actions[0]
        assert "Match Type" in kw_actions[0]

    def test_skips_when_all_effective(self):
        camps = [_camp(cid="100", spend=50, cpc=1.0)]
        placements = [
            _placement(cid="100", spend=30, sales=120, clicks=30),  # ACOS 25%
        ]
        result = _optimize_placements(camps, placements, {}, DEFAULTS)
        assert len(result) == 0

    def test_skips_when_ineffective_already_at_zero(self):
        """Skip suggestions when all ineffective placements already have 0% (0→0 no-op)."""
        camps = [_camp(cid="100", spend=50, cpc=1.0)]
        placements = [
            # Good placement at 100%
            _placement(cid="100", placement="Placement Top", percentage=100,
                       spend=30, sales=120, clicks=30),  # ACOS 25% - good
            # Bad placement already at 0% - nothing to reduce
            _placement(cid="100", placement="Placement Rest Of Search", percentage=0,
                       spend=20, sales=40, clicks=20),   # ACOS 50% - bad but already 0%
        ]
        result = _optimize_placements(camps, placements, {}, DEFAULTS)
        # Should not generate suggestion since ineffective placement is already 0%
        assert len(result) == 0

    def test_only_adjusts_nonzero_ineffective(self):
        """When some ineffective placements are 0% and others are not, only adjust the non-zero ones."""
        camps = [_camp(cid="100", spend=50, cpc=1.0)]
        placements = [
            # Good placement
            _placement(cid="100", placement="Placement Top", percentage=100,
                       spend=20, sales=100, clicks=20),  # ACOS 20% - good
            # Bad placement that CAN be reduced
            _placement(cid="100", placement="Placement Product Pages", percentage=50,
                       spend=15, sales=30, clicks=15),   # ACOS 50% - bad, can reduce
            # Bad placement already at 0% - skip this one
            _placement(cid="100", placement="Placement Rest Of Search", percentage=0,
                       spend=15, sales=30, clicks=15),   # ACOS 50% - bad but already 0%
        ]
        targets_by_camp = {"100": [
            _target(cid="100", keyword_text="kw1", match_type="Broad", bid=0.80),
        ]}
        result = _optimize_placements(camps, placements, targets_by_camp, DEFAULTS)
        # Should generate suggestion for Product Pages but not Rest Of Search
        assert len(result) == 1
        assert "Placement Product Pages" in result[0]["title"]
        # Rest Of Search should NOT be in the title (it's already at 0%)
        assert "Placement Rest Of Search" not in result[0]["title"]


class TestIncreaseBids:
    def test_suggests_increase_for_strong_campaign(self):
        camps = [_camp(
            spend=50, sales=500, orders=50, clicks=100, cpc=0.50,
        )]
        targets_by_camp = {"100": [
            _target(cid="100", keyword_text="kw1", match_type="Exact",
                    bid=0.50, spend=25, orders=25),
        ]}
        # CVR = 50%, ACOS = 10%
        result = _increase_bids(camps, [], targets_by_camp, DEFAULTS)
        assert len(result) == 1
        assert result[0]["category"] == "Increase Bids"
        assert result[0]["metrics"]["suggested_cpc"] > 0.50
        # Should include Keyword bid update row with increased Bid
        kw_actions = [a for a in result[0]["actions"] if a["Entity"] == "Keyword"]
        assert len(kw_actions) == 1
        assert kw_actions[0]["Bid"] > 0.50
        assert kw_actions[0]["Keyword Text"] == "kw1"
        assert kw_actions[0]["Match Type"] == "Exact"

    def test_skips_low_cvr(self):
        camps = [_camp(spend=50, sales=500, orders=5, clicks=100, cpc=0.50)]
        # CVR = 5% < 30%
        result = _increase_bids(camps, [], {}, DEFAULTS)
        assert len(result) == 0


class TestBulkXLSX:
    def test_generates_valid_xlsx(self):
        import io
        import openpyxl

        sugs = [{
            "id": "test_1",
            "actions": [
                {"Product": "Sponsored Products", "Entity": "Campaign",
                 "Operation": "Update", "Campaign ID": "123",
                 "State": "paused"},
            ],
        }]
        xlsx_bytes = build_bulk_xlsx(sugs)
        assert isinstance(xlsx_bytes, bytes)
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        # Header row
        headers = [ws.cell(row=1, column=c).value for c in range(1, 4)]
        assert headers == ["Product", "Entity", "Operation"]
        # Data row
        assert ws.cell(row=2, column=1).value == "Sponsored Products"
        assert ws.cell(row=2, column=2).value == "Campaign"
        assert ws.cell(row=2, column=3).value == "Update"
        wb.close()

    def test_empty_suggestions(self):
        import io
        import openpyxl

        xlsx_bytes = build_bulk_xlsx([])
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        # Only header row, no data
        assert ws.max_row == 1
        wb.close()


class TestCustomRules:
    def test_custom_rule_overrides_defaults_for_negatives(self):
        """Custom rules should override default thresholds."""
        # This term has only 5 clicks (below default threshold of 10)
        terms = [_search_term(term="low clicks", clicks=5, orders=0, spend=20, sales=0)]

        # Without custom rules, it should be skipped (clicks < 10)
        result = _negative_search_terms(terms, set(), DEFAULTS)
        assert len(result) == 0

        # With custom rule that triggers on clicks >= 3
        thresholds = {
            **DEFAULTS,
            "custom_rules": {
                "negatives": [
                    {
                        "id": "custom_1",
                        "enabled": True,
                        "conditions": [
                            {"metric": "clicks", "operator": ">=", "value": 3},
                            {"metric": "orders", "operator": "==", "value": 0},
                        ],
                        "action": {"type": "add_negative", "match_type": "Negative Phrase"},
                    }
                ],
            },
        }
        result = _negative_search_terms(terms, set(), thresholds)
        assert len(result) == 1
        assert result[0]["actions"][0]["Match Type"] == "Negative Phrase"

    def test_custom_rule_for_pause_campaigns(self):
        """Custom pause rules can trigger on different conditions."""
        # Campaign with low spend (below default threshold of 15) but high ACOS
        camps = [_camp(spend=8, orders=0, sales=0, clicks=100)]

        # Without custom rules, it should be skipped (spend < 15)
        result = _pause_campaigns(camps, DEFAULTS)
        assert len(result) == 0

        # With custom rule that triggers on spend >= 5
        thresholds = {
            **DEFAULTS,
            "custom_rules": {
                "pause": [
                    {
                        "id": "pause_1",
                        "enabled": True,
                        "conditions": [
                            {"metric": "spend", "operator": ">=", "value": 5},
                            {"metric": "orders", "operator": "==", "value": 0},
                        ],
                        "action": {"type": "pause"},
                    }
                ],
            },
        }
        result = _pause_campaigns(camps, thresholds)
        assert len(result) == 1

    def test_disabled_rule_is_ignored(self):
        """Disabled rules should not be applied."""
        terms = [_search_term(term="test", clicks=5, orders=0, spend=20, sales=0)]

        thresholds = {
            **DEFAULTS,
            "custom_rules": {
                "negatives": [
                    {
                        "id": "disabled_1",
                        "enabled": False,  # Disabled
                        "conditions": [
                            {"metric": "clicks", "operator": ">=", "value": 1},
                        ],
                        "action": {"type": "add_negative"},
                    }
                ],
            },
        }
        result = _negative_search_terms(terms, set(), thresholds)
        # Should be empty because rule is disabled and default threshold not met
        assert len(result) == 0


class TestGenerateAll:
    def test_full_pipeline(self):
        analysis = {
            "campaigns_table": [
                _camp(cid="1", spend=25, orders=0, sales=0),
                _camp(cid="2", spend=50, sales=500, orders=50, clicks=100, cpc=0.5),
            ],
            "targets": [
                _target(cid="1", spend=15, orders=0),
            ],
            "placements": [],
            "search_terms_detail": [
                _search_term(term="wasted", clicks=20, orders=0, spend=15, sales=0,
                             source_type="broad"),
            ],
            "existing_exact_keywords": [],
            "existing_negatives": [],
            "product_ads": [],
            "portfolios": [],
        }
        sugs = generate_suggestions(analysis)
        assert len(sugs) > 0
        categories = {s["category"] for s in sugs}
        assert "Pause Campaigns" in categories
        assert "Search Term Negatives" in categories
        # All suggestions have IDs
        assert all(s.get("id") for s in sugs)


if __name__ == "__main__":
    import pytest
    pytest.main([__file__, "-v"])
